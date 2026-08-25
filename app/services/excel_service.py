from pathlib import Path
from openpyxl import load_workbook
from app.models.job import Job
from datetime import datetime


TEMPLATE_PATH = Path("templates/Job_Tracker_Template.xlsx")
SYSTEM_DATE = "%Y%m%d%H%M%S"


def write_jobs_to_excel(jobs: list[Job]):

    # Open template
    workbook = load_workbook(TEMPLATE_PATH)

    # Get first sheet
    worksheet = workbook.active

    # Start write from row index
    start_row = 9
    # STT
    stt = 1

    for index, job in enumerate(jobs, start=start_row):
        stt = index - start_row + 1

        worksheet.cell(index, 1, stt)
        worksheet.cell(index, 2, job.title)
        worksheet.cell(index, 3, job.company)
        worksheet.cell(index, 4, job.salary)
        worksheet.cell(index, 5, job.location)
        worksheet.cell(index, 6, job.experience)
        worksheet.cell(index, 7, job.source)
        worksheet.cell(index, 8, job.status)
        worksheet.cell(index, 9, job.url)

    # System date
    sys_date = datetime.now().strftime("%Y%m%d%H%M%S")

    # Path object
    output_path = Path(f"output/jobs_{sys_date}.xlsx")

    # Create output folder
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Save
    workbook.save(output_path)
    print(f"Excel saved: {output_path}")
